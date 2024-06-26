import time
import io
import re
import pickle
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from flask import Flask, render_template, request, session, send_file
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys

# Configuração do aplicativo Flask
app = Flask(__name__)
app.secret_key = 'dato123'  # Defina uma chave secreta adequada

# Configuração do Selenium
from webdriver_manager.chrome import ChromeDriverManager
servico = Service(ChromeDriverManager().install())
options = Options()
options.add_argument("--no-sandbox")
options.add_argument("--headless")  # Executar o Chrome de forma oculta
options.add_argument("--disable-dev-shm-usage")
driver = webdriver.Chrome(options=options)

# Inicializa o driver (supondo que você já tenha o ChromeDriver configurado)
wait = WebDriverWait(driver, 30)

# Variáveis de e-mail e senha
lista_pendentes = []
email = 'daniele@amplologistica.com.br'
senha = 'buslog@2021'
is_first_execution = True
# Variável global para armazenar o DataFrame atualizado
df_atualizado = None

@app.route('/')
def index():
    return render_template('index.html', pendentes=[])

@app.route('/resultado')
def resultado():
    global df_atualizado
    statuses, datas, df = capturar_status_pendentes()
    df_atualizado = df  # Atualiza o DataFrame global
    table_html = df.to_html(classes='table table-bordered', index=False)
    return render_template('resultado.html', table_html=table_html)

@app.route('/exportar_excel')
def exportar_excel():
    global df_atualizado
    if df_atualizado is None:
        return "Nenhum dado disponível para exportação"

    # Utilizando um buffer em memória para salvar o DataFrame como Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_atualizado.to_excel(writer, index=False, sheet_name='Dados_Buslog')
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name='ResultadoBuslog.xlsx', as_attachment=True)

@app.route('/', methods=['POST'])
def preparar_dados_planilha():
    global lista_pendentes  # Acessando a variável global
    file = request.files['file']

    if not file.filename.endswith('.xlsx'):
        return "Por favor, selecione um arquivo Excel (.xlsx)"

    # Salvar o arquivo Excel em uma pasta temporária
    filename = secure_filename(file.filename)
    file.save(filename)
    session['excel_filename'] = filename  # Armazenar o nome do arquivo na sessão

    planilha = load_workbook(filename)
    aba_ativa = planilha.active

    lista_pendentes = []
    for coluna_a, coluna_c, coluna_d in zip(aba_ativa["A"][1:], aba_ativa["C"][1:], aba_ativa["D"][1:]):
        if coluna_d.value != 'ENTREGUE':
            if coluna_c.value is not None:
                lista_pendentes.append(coluna_c.value)
    total_pendentes = len(lista_pendentes)
    print("=" * 150)
    print(f'As pendente de entrega são: Total: |{total_pendentes}| {lista_pendentes}')
    print("=" * 150)

    return render_template('index.html', total_pendentes=total_pendentes, pendentes=lista_pendentes)

def login(driver, wait, email, senha):
    try:
        driver.get(f"https://www.track3r.com.br/tms/default")
        # Localiza o campo de e-mail pelo XPath e insere o e-mail
        campo_email = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="txtEmail"]')))
        campo_email.send_keys(email)
        # Envia a tecla ENTER para confirmar o e-mail
        campo_email.send_keys(Keys.ENTER)

        # Localiza o campo de senha pelo XPath e insere a senha
        campo_senha = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="txtSenha"]')))
        campo_senha.send_keys(senha)
        # Envia a tecla ENTER para confirmar a senha
        campo_senha.send_keys(Keys.ENTER)

        

    except TimeoutException as e:
        print(f"Erro durante o login: {e}")


def captura_status(i, cod_buslog, awb):
    global is_first_execution

    try:
        # Se for a primeira execução, realiza o login
        if is_first_execution:
            login(driver, wait, email, senha)
            is_first_execution = False
            
     
            # Recarrega a página de tracking após o login
            driver.get(f"https://www.track3r.com.br/tms/consultas-tracking?Encomenda={awb}&TipoBusca=2")

        # Abre a página de tracking com o número de encomenda especificado
        driver.get(f"https://www.track3r.com.br/tms/consultas-tracking?Encomenda={awb}&TipoBusca=2")

        coleta_status = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_ucTrackingEncomenda_gdvRegistros_ctl02_lblDescStatus"]')))
        status = coleta_status.text

        coleta_data = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="ContentPlaceHolder1_ucTrackingEncomenda_gdvRegistros"]/tbody/tr[2]/td[1]')))
        data = coleta_data.text


        campo_envolvido = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="ContentPlaceHolder1_ucTrackingEncomenda_liEnvolvidos"]/a')))
        campo_envolvido.click()
        

        nome_destinatario = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="lblNmDestinatario"]')))
        destinatario = nome_destinatario.text

        cnpj_destinatario = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="lblNrCpfCnpjDestinatario"]')))
        cnpj = cnpj_destinatario.text
        cnpj_formatado = re.sub(r'[\.\-\/]', '', cnpj)

        # Suponho que `lista_pendentes` seja definida em algum lugar do seu código
        total = len(lista_pendentes) if 'lista_pendentes' in globals() else 0
        
        
        captura = f'|TOTAL: {i+1}/{total} |COD BUSLOG: {cod_buslog}|AWB: {awb}| STATUS: {status}|DATA_EVENTO: {data}|'
        print("=" * 150)
        print(captura)
        print("=" * 150)

        return destinatario, cnpj_formatado, status, data
    except TimeoutException:
        destinatario = "Erro de tempo limite (Buslog indisponível)"
        cnpj = "Erro de tempo limite (Buslog indisponível)"
        status = "Erro de tempo limite (Buslog indisponível)"
        data = "Erro de tempo limite (Buslog indisponível)"
        print("Erro de tempo limite (Buslog indisponível)")
        return destinatario, cnpj_formatado, status, data

def capturar_status_pendentes():
    dados_rastreamento = []
    destinatarioes = []
    cnpjes = []

    # Obter o nome do arquivo Excel da sessão
    excel_filename = session.get('excel_filename')

    if excel_filename:
        planilha = load_workbook(excel_filename)
        aba_ativa = planilha.active

        for i, (coluna_a, coluna_c, coluna_d) in enumerate(zip(aba_ativa["A"][1:], aba_ativa["C"][1:], aba_ativa["D"][1:])):
            if coluna_d.value != 'ENTREGUE':
                if coluna_c.value is not None:
                    cod_buslog = coluna_a.value
                    awb = coluna_c.value
                    status = coluna_d.value
                    destinatario, cnpj, status, data = captura_status(i, cod_buslog, awb)
                    dados_rastreamento.append({
                        'COD BUSLOG': cod_buslog,
                        'AWB': awb,
                        'STATUS': status,
                        'DATA_EVENTO': data,
                        'DESTINAÁRIO': destinatario,
                        'CNPJ': cnpj

                    })
                    destinatarioes.append(destinatario)
                    cnpjes.append(cnpj)

        df = pd.DataFrame(dados_rastreamento)
        print(f'captura_status_pendentes::::  {df}')

        c = pd.DataFrame(dados_rastreamento)
        print(f'captura_status_pendentes df_exportar::::  {df}')
        return destinatarioes, cnpjes, df
    else:
        return [], [], None  # Retorna vazios se o arquivo não estiver definido na sessão

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=9090)
